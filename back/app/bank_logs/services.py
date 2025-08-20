# app/bank_logs/services.py
from __future__ import annotations
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List
import io
from openpyxl import Workbook
from openpyxl.styles import Font

from sqlalchemy.orm import joinedload
from sqlalchemy.exc import IntegrityError

from app import db
from app.base_service import BaseService
from app.errors import AppError
from app.logging_decorator import service_logger
from app.logging_utils import dinfo, dwarn, derr
from app.banks.models import BankAccount, Bank
from .models import BankLog, Period
from ..banks.schemas import BankSchema


# ------------------------- helpers -------------------------

def _coerce_date(v: Any, field: str = "date") -> date:
    """
    'YYYY-MM-DD' bekler. Hatalıysa AppError(400).
    (İstersen '%d.%m.%Y' desteği de ekleyebilirsin.)
    """
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, str):
        try:
            return datetime.strptime(v.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise AppError(f"{field} must be YYYY-MM-DD.", 400, code="INVALID_DATE", details={"field": field, "given": v})
    raise AppError(f"{field} has invalid type.", 400, code="INVALID_DATE_TYPE", details={"field": field, "type": type(v).__name__})

def _coerce_period(v: Any, field: str = "period") -> Period:
    """
    'morning' / 'evening' (value) veya 'MORNING' / 'EVENING' (name) ya da 'Period.MORNING' gibi şeyleri kabul et.
    Hatalıysa AppError(400).
    """
    if isinstance(v, Period):
        return v
    if isinstance(v, str):
        s = v.strip()
        # 'Period.MORNING' -> 'MORNING'
        if "." in s:
            s = s.split(".")[-1]
        # value olarak ('morning') gelirse:
        lowered = s.lower()
        for p in Period:
            if getattr(p, "value", "").lower() == lowered:
                return p
        # name olarak ('MORNING') gelirse:
        uppered = s.upper()
        try:
            return Period[uppered]  # type: ignore[index]
        except Exception:
            pass
    raise AppError(f"{field} is invalid.", 400, code="INVALID_PERIOD", details={"given": v})

def _to_decimal(v: Any, field: str) -> Decimal | None:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        raise AppError(f"{field} must be a number.", 400, code="INVALID_DECIMAL", details={"field": field, "given": v})

def _serialize_amount(x: Decimal | None) -> str | None:
    return None if x is None else f"{x}"


# ------------------------- service -------------------------

class BankLogService(BaseService):
    def __init__(self):
        super().__init__(BankLog)

    # ------- READ -------

    @service_logger
    def get_all_logs_for_period(self, date_str: str, period_str: str) -> List[Dict[str, Any]]:
        """
        Verilen gün + period için tüm bankaların loglarını döner.
        Log olmayan bankalar için placeholder üretir.
        """
        d = _coerce_date(date_str, "date")
        p = _coerce_period(period_str, "period")

        banks = Bank.query.options(
            joinedload(Bank.accounts)  # gerekirse
        ).all()

        bank_schema = BankSchema()
        out: List[Dict[str, Any]] = []

        for bank in banks:
            log = (
                self.model.query
                .filter_by(bank_id=bank.id, date=d, period=p)
                .first()
            )

            bank_data = bank_schema.dump(bank)

            if log:
                bank_data["log"] = {
                    "id": log.id,
                    "bank_id": log.bank_id,
                    "date": d.isoformat(),
                    "period": p.value if hasattr(p, "value") else str(p),
                    "amount_try": _serialize_amount(log.amount_try),
                    "amount_usd": _serialize_amount(log.amount_usd),
                    "amount_eur": _serialize_amount(log.amount_eur),
                    "amount_aed": _serialize_amount(log.amount_aed),
                    "amount_gbp": _serialize_amount(log.amount_gbp),
                    "rate_usd_try": _serialize_amount(log.rate_usd_try),
                    "rate_eur_try": _serialize_amount(log.rate_eur_try),
                    "rate_aed_try": _serialize_amount(log.rate_aed_try),
                    "rate_gbp_try": _serialize_amount(log.rate_gbp_try),
                }
            else:
                bank_data["log"] = {
                    "id": f"new-{bank.id}-{d.isoformat()}-{p.name}",
                    "bank_id": bank.id,
                    "date": d.isoformat(),
                    "period": p.value if hasattr(p, "value") else str(p),
                    "amount_try": "0.00",
                    "amount_usd": "0.00",
                    "amount_eur": "0.00",
                    "amount_aed": "0.00",
                    "amount_gbp": "0.00",
                    "rate_usd_try": None,
                    "rate_eur_try": None,
                    "rate_aed_try": None,
                    "rate_gbp_try": None,
                }

            out.append(bank_data)

        dinfo("bank_logs.by_period", date=d.isoformat(), period=p.name, banks=len(banks), rows=len(out))
        return out

    # ------- UPSERT (single/batch) -------

    def _prepare_log_from_data(self, data: Dict[str, Any], existing: BankLog | None = None) -> BankLog:
        """Body'yi parse eder ve (varsa) mevcut logu patch'ler; yoksa yeni instance döner (session'a ekleme yapmaz)."""
        if not isinstance(data, dict):
            raise AppError("Body must be an object.", 400, code="INVALID_BODY")

        required = ["bank_id", "date", "period", "amount_try", "amount_usd", "amount_eur", "amount_aed", "amount_gbp"]
        missing = [k for k in required if k not in data]
        if missing:
            raise AppError("Missing required fields.", 400, code="MISSING_FIELDS", details={"fields": missing})

        bank_id = int(data["bank_id"])
        d = _coerce_date(data["date"], "date")
        p = _coerce_period(data["period"], "period")

        # miktarlar
        attrs = {
            "amount_try": _to_decimal(data.get("amount_try"), "amount_try"),
            "amount_usd": _to_decimal(data.get("amount_usd"), "amount_usd"),
            "amount_eur": _to_decimal(data.get("amount_eur"), "amount_eur"),
            "amount_aed": _to_decimal(data.get("amount_aed"), "amount_aed"),
            "amount_gbp": _to_decimal(data.get("amount_gbp"), "amount_gbp"),
            "rate_usd_try": _to_decimal(data.get("rate_usd_try"), "rate_usd_try"),
            "rate_eur_try": _to_decimal(data.get("rate_eur_try"), "rate_eur_try"),
            "rate_aed_try": _to_decimal(data.get("rate_aed_try"), "rate_aed_try"),
            "rate_gbp_try": _to_decimal(data.get("rate_gbp_try"), "rate_gbp_try"),
        }

        # mevcut var mı?
        log = existing or (
            self.model.query
            .filter_by(bank_id=bank_id, date=d, period=p)
            .first()
        )

        if log:
            for k, v in attrs.items():
                setattr(log, k, v)
        else:
            log = self.model(bank_id=bank_id, date=d, period=p, **attrs)

        return log

    @service_logger
    def create_or_update_log(self, data: Dict[str, Any], commit: bool = True) -> BankLog:
        log = self._prepare_log_from_data(data)
        if not getattr(log, "id", None):
            db.session.add(log)

        if commit:
            try:
                db.session.commit()
            except IntegrityError as ie:
                db.session.rollback()
                # muhtemel uniq violation vs.
                raise AppError("Could not save bank log (integrity).", 409, code="INTEGRITY_ERROR")
            except Exception as ex:
                db.session.rollback()
                derr("bank_logs.upsert.unhandled", err=ex)
                raise

        dinfo("bank_logs.upsert", bank_id=log.bank_id, date=log.date.isoformat(), period=log.period.name)
        return log

    @service_logger
    def batch_upsert_logs(self, logs_data: List[Dict[str, Any]]) -> List[BankLog]:
        if not isinstance(logs_data, list) or not logs_data:
            raise AppError("Body must be a non-empty list.", 400, code="INVALID_BODY")

        rows: List[BankLog] = []
        for item in logs_data:
            log = self._prepare_log_from_data(item)
            if not getattr(log, "id", None):
                db.session.add(log)
            rows.append(log)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            raise AppError("Could not save bank logs (integrity).", 409, code="INTEGRITY_ERROR")
        except Exception as ex:
            db.session.rollback()
            derr("bank_logs.batch.unhandled", err=ex)
            raise

        dinfo("bank_logs.batch_upsert", count=len(rows))
        return rows

    @service_logger
    def generate_balance_excel(self, date_str):
        """
        Verilen tarihe göre banka bakiyelerini ve kurları alıp
        SABAH ve AKŞAM toplamlarını ayrı ayrı gösteren formatta Excel dosyası oluşturur.
        """
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("Geçersiz tarih formatı. Lütfen YYYY-MM-DD formatını kullanın.")

        logs = self.model.query.options(joinedload(BankLog.bank)).filter_by(date=target_date).order_by(BankLog.bank_id, BankLog.period).all()

        if not logs:
            raise ValueError("Seçilen tarih için görüntülenecek veri bulunamadı.")

        wb = Workbook()
        ws = wb.active
        ws.title = f"Bakiye Raporu {date_str}"

        # --- Başlık ve Stil Ayarları ---
        header_font = Font(bold=True)
        
        # --- Kur Bilgilerini Hazırlama ---
        first_valid_log = next((log for log in logs if log.rate_usd_try is not None), logs[0])
        rates = {
            'usd': first_valid_log.rate_usd_try or Decimal(0),
            'eur': first_valid_log.rate_eur_try or Decimal(0),
            'aed': first_valid_log.rate_aed_try or Decimal(0),
            'gbp': first_valid_log.rate_gbp_try or Decimal(0),
        }

        # --- YENİ: Sabah ve Akşam için Ayrı Ayrı Genel Toplam Hesaplamaları ---
        def get_period_totals(period_logs):
            totals = {
                'try': Decimal(0), 'usd': Decimal(0), 'eur': Decimal(0),
                'aed': Decimal(0), 'gbp': Decimal(0), 'in_try': Decimal(0)
            }
            for log in period_logs:
                totals['try'] += log.amount_try or Decimal(0)
                totals['usd'] += log.amount_usd or Decimal(0)
                totals['eur'] += log.amount_eur or Decimal(0)
                totals['aed'] += log.amount_aed or Decimal(0)
                totals['gbp'] += log.amount_gbp or Decimal(0)
            
            totals['in_try'] = (totals['try'] + 
                                (totals['usd'] * rates['usd']) +
                                (totals['eur'] * rates['eur']) +
                                (totals['aed'] * rates['aed']) +
                                (totals['gbp'] * rates['gbp']))
            return totals

        morning_grand_totals = get_period_totals([log for log in logs if log.period == Period.morning])
        evening_grand_totals = get_period_totals([log for log in logs if log.period == Period.evening])

        # --- Excel İçeriğini Oluşturma (YENİ FORMATA GÖRE) ---

        # Satır 1: Tarih Bilgisi
        ws['A1'] = "tarih"
        ws['B1'] = target_date.strftime('%d.%m.%Y')
        ws['A1'].font = header_font
        
        # Satır 3: TOPLAM BAKİYE (SABAH)
        ws.cell(row=5, column=1, value="TOPLAM BAKİYE").font = header_font
        ws.cell(row=5, column=2, value="SABAH").font = header_font
        ws.cell(row=5, column=3, value=morning_grand_totals['in_try'])
        ws.cell(row=5, column=4, value=morning_grand_totals['try'])
        ws.cell(row=5, column=5, value=morning_grand_totals['usd'])
        ws.cell(row=5, column=6, value=morning_grand_totals['eur'])
        ws.cell(row=5, column=7, value=morning_grand_totals['aed'])
        ws.cell(row=5, column=8, value=morning_grand_totals['gbp'])

        # Satır 4: TOPLAM BAKİYE (AKŞAM)
        ws.cell(row=6, column=1, value="TOPLAM BAKİYE").font = header_font
        ws.cell(row=6, column=2, value="AKŞAM").font = header_font
        ws.cell(row=6, column=3, value=evening_grand_totals['in_try'])
        ws.cell(row=6, column=4, value=evening_grand_totals['try'])
        ws.cell(row=6, column=5, value=evening_grand_totals['usd'])
        ws.cell(row=6, column=6, value=evening_grand_totals['eur'])
        ws.cell(row=6, column=7, value=evening_grand_totals['aed'])
        ws.cell(row=6, column=8, value=evening_grand_totals['gbp'])
            
        # Ana Tablo Başlıkları (Satır 6)
        ws.cell(row=3, column=1, value="Banka").font = header_font
        ws.cell(row=3, column=2, value="Vakit").font = header_font
        ws.cell(row=3, column=3, value="Toplam (TRY)").font = header_font
        ws.cell(row=3, column=4, value="TRY").font = header_font
        ws.cell(row=3, column=5, value="USD").font = header_font
        ws.cell(row=3, column=6, value="EUR").font = header_font
        ws.cell(row=3, column=7, value="AED").font = header_font
        ws.cell(row=3, column=8, value="GBP").font = header_font

        # Veri Satırlarını Doldurma
        def calculate_row_total(log):
            if not log: return Decimal(0)
            return ((log.amount_try or Decimal(0)) +
                    (log.amount_usd or Decimal(0)) * rates['usd'] +
                    (log.amount_eur or Decimal(0)) * rates['eur'] +
                    (log.amount_aed or Decimal(0)) * rates['aed'] +
                    (log.amount_gbp or Decimal(0)) * rates['gbp'])

        current_row = 7
        banks = sorted(list(set([log.bank for log in logs])), key=lambda b: b.id)
        for bank in banks:
            # Sabah Kaydı
            morning_log = next((log for log in logs if log.bank_id == bank.id and log.period == Period.morning), None)
            ws.cell(row=current_row, column=1, value=bank.name)
            ws.cell(row=current_row, column=2, value="sabah")
            ws.cell(row=current_row, column=3, value=calculate_row_total(morning_log))
            ws.cell(row=current_row, column=4, value=morning_log.amount_try if morning_log else 0)
            ws.cell(row=current_row, column=5, value=morning_log.amount_usd if morning_log else 0)
            ws.cell(row=current_row, column=6, value=morning_log.amount_eur if morning_log else 0)
            ws.cell(row=current_row, column=7, value=morning_log.amount_aed if morning_log else 0)
            ws.cell(row=current_row, column=8, value=morning_log.amount_gbp if morning_log else 0)
            current_row += 1

            # Akşam Kaydı
            evening_log = next((log for log in logs if log.bank_id == bank.id and log.period == Period.evening), None)
            ws.cell(row=current_row, column=1, value=bank.name)
            ws.cell(row=current_row, column=2, value="akşam")
            ws.cell(row=current_row, column=3, value=calculate_row_total(evening_log))
            ws.cell(row=current_row, column=4, value=evening_log.amount_try if evening_log else 0)
            ws.cell(row=current_row, column=5, value=evening_log.amount_usd if evening_log else 0)
            ws.cell(row=current_row, column=6, value=evening_log.amount_eur if evening_log else 0)
            ws.cell(row=current_row, column=7, value=evening_log.amount_aed if evening_log else 0)
            ws.cell(row=current_row, column=8, value=evening_log.amount_gbp if evening_log else 0)
            current_row += 1

        # Kurlar Bölümü
        ws.cell(row=6, column=10, value="Kurlar").font = header_font
        ws.cell(row=7, column=10, value="USD/TRY")
        ws.cell(row=8, column=10, value="EUR/TRY")
        ws.cell(row=9, column=10, value="AED/TRY")
        ws.cell(row=10, column=10, value="GBP/TRY")

        ws.cell(row=7, column=11, value=rates['usd'])
        ws.cell(row=8, column=11, value=rates['eur'])
        ws.cell(row=9, column=11, value=rates['aed'])
        ws.cell(row=10, column=11, value=rates['gbp'])

        # Sayı Formatlarını ve Sütun Genişliklerini Ayarla
        money_format = '#,##0.00'
        rate_format = '#,##0.0000'

        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(3, 9): # Sütun C'den H'ye kadar
                ws.cell(row=row_idx, column=col_idx).number_format = money_format
        
        for row_idx in range(7, 11): # Kur formatları
            ws.cell(row=row_idx, column=11).number_format = rate_format

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].autosize = True

        # Excel dosyasını hafızada bir stream'e kaydet
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)

        return virtual_workbook      

bank_log_service = BankLogService()